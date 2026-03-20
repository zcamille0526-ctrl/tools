#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
教材知识点梳理与旧库去重工具 v2.0
==================================
使用方式：python3 app.py
然后浏览器打开 http://localhost:8788
"""

import os
import sys
import json
import re
import html as html_mod
import time
import uuid
import threading
import traceback
import urllib.request
import urllib.parse
from http.server import HTTPServer, BaseHTTPRequestHandler
from socketserver import ThreadingMixIn
from pathlib import Path
from difflib import SequenceMatcher

# ============================================================
# 配置
# ============================================================
PORT = 8788
BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# ============================================================
# 百度OCR配置（扫描版PDF专用）
# ============================================================
BAIDU_APP_ID     = os.environ.get("BAIDU_APP_ID", "122411799")
BAIDU_API_KEY    = os.environ.get("BAIDU_API_KEY", "tShoflxPjqFfUdQI0xjuTQuo")
BAIDU_SECRET_KEY = os.environ.get("BAIDU_SECRET_KEY", "kCWeOSxvxueIIShFwWDUy5xLzR3KlgEA")  # ⚠️ 建议通过环境变量设置，勿提交到版本控制
POPPLER_PATH     = os.environ.get("POPPLER_PATH", r"D:\poppler\Library\bin")

# ============================================================
# PDF文本提取（自动识别电子版/扫描版）
# ============================================================
def extract_pdf_text(pdf_path):
    """从PDF中提取文本，自动判断电子版/扫描版，扫描版走百度OCR"""
    import pdfplumber

    # ── 第一步：先尝试直接提取文字（电子版PDF）──
    pages = {}
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            pages[i + 1] = text

    total_chars = sum(len(t) for t in pages.values())

    # ── 提取到足够文字，直接返回（电子版）──
    if total_chars > 100:
        return pages

    # ── 文字太少，说明是扫描版，启动百度OCR ──
    print("⚠️  检测到扫描版PDF，启动百度OCR识别...")

    try:
        from aip import AipOcr
        from pdf2image import convert_from_path
        import base64, io

        client = AipOcr(BAIDU_APP_ID, BAIDU_API_KEY, BAIDU_SECRET_KEY)

        with pdfplumber.open(pdf_path) as pdf:
            total = len(pdf.pages)

        ocr_pages = {}
        BATCH = 10  # 每批10页，防止内存溢出

        for batch_start in range(0, total, BATCH):
            batch_end = min(batch_start + BATCH, total)
            print(f"  OCR处理第 {batch_start+1}~{batch_end} 页 / 共{total}页...")

            images = convert_from_path(
                pdf_path,
                first_page=batch_start + 1,
                last_page=batch_end,
                dpi=150,              # 150dpi够用，更高会超百度图片大小限制4MB
                poppler_path=POPPLER_PATH
            )

            for j, img in enumerate(images):
                page_num = batch_start + j + 1

                # 图片转base64发给百度
                buf = io.BytesIO()
                img.save(buf, format="PNG")
                img_b64 = base64.b64encode(buf.getvalue()).decode()

                # 调用百度高精度OCR
                result = client.basicAccurate(base64.b64decode(img_b64))

                if "words_result" in result:
                    text = "\n".join(
                        w["words"] for w in result["words_result"]
                    )
                else:
                    text = ""
                    print(f"  ⚠️  第{page_num}页识别失败：{result.get('error_msg', '未知错误')}")

                ocr_pages[page_num] = text

                # 百度免费版限速：每秒2次，加延迟避免超限报错
                time.sleep(0.6)

        print(f"✅ OCR完成，共识别 {total} 页")
        return ocr_pages

    except ImportError as e:
        print(f"❌ 缺少依赖库：{e}")
        print("请运行：pip install baidu-aip pdf2image Pillow")
        return pages


# ============================================================
# 目录解析与章节分割（改进版）
# ============================================================
def _match_toc_entry(line, prefix_pattern):
    """匹配目录行，同时支持省略号格式（第X章 标题…3）和空格格式（第X章 标题 3）
    以及PDF渲染artifact中用!作分隔符的格式（第 15 章 ! 分式!1）
    返回 (prefix, title, page) 或 None"""
    m = re.search(rf"({prefix_pattern})\s*(.+?)…+\s*(\d+)", line)
    if m:
        return m.group(1), m.group(2).strip(), int(m.group(3))
    m = re.search(rf"^({prefix_pattern})\s+(.+?)\s+(\d+)$", line)
    if m:
        return m.group(1), m.group(2).strip(), int(m.group(3))
    # 兼容PDF提取artifact：! 作为分隔符，如"第 15 章 ! 分式!1"
    m = re.search(rf"^({prefix_pattern})[!\s]+(.+?)[!\s]+(\d+)[!\s]*$", line)
    if m:
        title = re.sub(r'[!]+', ' ', m.group(2)).strip()
        if title:
            return m.group(1), title, int(m.group(3))
    return None


def parse_toc_entries(pages_dict):
    """从PDF页面中解析目录，返回结构化的目录条目列表"""
    toc_pages_text = ""
    for pnum in sorted(pages_dict.keys()):
        text = pages_dict[pnum]
        if "目录" in text or "目 录" in text:
            toc_pages_text += text + "\n"
            for next_p in range(pnum + 1, pnum + 4):
                if next_p in pages_dict:
                    next_text = pages_dict[next_p]
                    # 支持有省略号/空格格式、阿拉伯数字章节（第 15 章）以及小数节号（15"1 或 15.1）
                    if re.search(r"(…+\s*\d+|第[一二三四五六七八九十百]+[章节]\s+.+?\s+\d+|第\s*\d+\s*[章节]|\d+[\"\.]\d+)", next_text):
                        toc_pages_text += next_text + "\n"
                    else:
                        break
            break

    if not toc_pages_text:
        return []

    entries = []
    current_chapter = ""
    pending_theme_title = None  # 用于处理"主题探究"后续行

    for line in toc_pages_text.split("\n"):
        line = line.strip()
        if not line:
            continue

        # 处理"主题探究"后续行（带页码的标题）
        if pending_theme_title:
            # 从行尾提取页码
            page_num_match = re.search(r'(\d+)\s*$', line)
            if page_num_match:
                page_num = int(page_num_match.group(1))
                # 截掉页码及前面的省略号/空格，得到标题
                raw = line[:page_num_match.start()].strip()
                raw = re.sub(r'[…\s]+$', '', raw)  # 去除尾部省略号
                clean_title = re.sub(r'["""\']', '', raw).strip()  # 去除各类引号
                if clean_title:
                    entries.append({
                        "type": "section",
                        "chapter": current_chapter,
                        "section": f"主题探究 {clean_title}",
                        "title": clean_title,
                        "page": page_num,
                    })
                pending_theme_title = None
                continue
            else:
                pending_theme_title = None

        # 匹配"主题探究"（单独一行，无页码）
        if line == "主题探究":
            pending_theme_title = True
            continue

        # === 有省略号的格式（原有逻辑） ===

        # 匹配简单数字章节（一、二、三...）
        simple_ch_match = re.search(r"^([一二三四五六七八九十]+)[、\s]+(.+?)…+\s*(\d+)", line)
        if simple_ch_match:
            current_chapter = f"{simple_ch_match.group(1)} {simple_ch_match.group(2).strip()}"
            entries.append({
                "type": "chapter",
                "chapter": current_chapter,
                "section": "",
                "title": simple_ch_match.group(2).strip(),
                "page": int(simple_ch_match.group(3)),
            })
            continue

        # 匹配无数字前缀的章节（如"年、月、日的奥秘"）
        no_prefix_match = re.search(r"^([^一二三四五六七八九十第\d\s].{2,30}?)…+\s*(\d+)$", line)
        if no_prefix_match:
            title = no_prefix_match.group(1).strip()
            if not re.match(r"^\d+[\.、]", title) and len(entries) < 15:
                current_chapter = title
                entries.append({
                    "type": "chapter",
                    "chapter": current_chapter,
                    "section": "",
                    "title": title,
                    "page": int(no_prefix_match.group(2)),
                })
                continue

        # 章标题（省略号/空格两种格式统一处理）
        m = _match_toc_entry(line, r"第[一二三四五六七八九十百]+章")
        if m:
            prefix, title, page = m
            current_chapter = f"{prefix} {title}"
            entries.append({"type": "chapter", "chapter": current_chapter, "section": "", "title": title, "page": page})
            continue

        # 阿拉伯数字章标题（如"第 15 章 分式"）
        m = _match_toc_entry(line, r"第\s*\d+\s*章")
        if m:
            prefix, title, page = m
            prefix_clean = re.sub(r'\s+', '', prefix)  # "第15章"
            # 清理PDF artifact字符：% → 、，! → 空格
            title_clean = re.sub(r'%', '、', title)
            title_clean = re.sub(r'[!]+', ' ', title_clean).strip()
            title_clean = re.sub(r'、\s+', '、', title_clean)  # 清理顿号后多余空格
            current_chapter = f"{prefix_clean} {title_clean}"
            entries.append({"type": "chapter", "chapter": current_chapter, "section": "", "title": title_clean, "page": page})
            continue

        # 小数点节号（数学/理科教材）：支持"15.1 标题 2"和PDF artifact"15"1! 标题!2"
        # 也跳过纯数字小节（如"1"分式"2"）只取X.Y格式的主节
        m = re.search(r'^(\d{2,}["\.](\d+))[!\s]+(.{2,40}?)[!\s]+(\d+)[!\s]*$', line)
        if not m:
            m = re.search(r'^(\d{2,}\.(\d+))\s+(.{2,40}?)\s+(\d+)$', line)
        if m and current_chapter:
            raw_sec = m.group(1).replace('"', '.')  # 15"1 → 15.1
            sub_idx = int(m.group(2))
            title = re.sub(r'[!%]+', ' ', m.group(3)).strip()
            page = int(m.group(4))
            # 只保留顶层小节（15.1, 15.2...），跳过子子节（内嵌的 1., 2. 等单数字）
            if title and len(title) >= 2 and len(raw_sec) >= 4:
                entries.append({"type": "section", "chapter": current_chapter,
                                "section": f"{raw_sec} {title}", "title": title, "page": page})
                continue

        # 单元标题
        m = _match_toc_entry(line, r"第[一二三四五六七八九十百]+单元")
        if m:
            prefix, title, page = m
            current_chapter = f"{prefix} {title}"
            entries.append({"type": "chapter", "chapter": current_chapter, "section": "", "title": title, "page": page})
            continue

        # 节标题
        m = _match_toc_entry(line, r"第[一二三四五六七八九十百]+节")
        if m:
            prefix, title, page = m
            entries.append({"type": "section", "chapter": current_chapter, "section": f"{prefix} {title}", "title": title, "page": page})
            continue

        # 课标题（含活动课识别）
        m = _match_toc_entry(line, r"第\s*\d+\s*课")
        if m:
            prefix, title, page = m
            section_name = f"{prefix.replace(' ', '')} {title}"
            is_activity = "活动课" in prefix or "活动课" in title
            entries.append({"type": "skip" if is_activity else "section", "chapter": current_chapter, "section": section_name, "title": title, "page": page})
            continue

        # 跳过项（有省略号）
        skip_match = re.search(r"(【.+?】|跨学科|附录|本书常用|学史方法|活动课|大事年表).+?…+\s*(\d+)", line)
        if skip_match:
            entries.append({
                "type": "skip",
                "chapter": "",
                "section": skip_match.group(0).split("…")[0].strip(),
                "title": skip_match.group(0).split("…")[0].strip(),
                "page": int(skip_match.group(2)),
            })

    # 添加顺序索引
    for i, entry in enumerate(entries):
        entry["order_index"] = i

    return entries


def split_pages_by_toc(pages_dict, toc_entries):
    """根据目录条目把页面文本分配到各章节"""
    if not toc_entries:
        return split_pages_by_regex(pages_dict)

    # 校准PDF页码偏移
    offset = 0
    for entry in toc_entries:
        if entry["type"] in ("chapter", "section"):
            target_title = entry["title"][:6]
            target_page = entry["page"]
            for pdf_page in sorted(pages_dict.keys()):
                if target_title in pages_dict[pdf_page]:
                    offset = pdf_page - target_page
                    break
            break

    for entry in toc_entries:
        entry["pdf_page"] = entry["page"] + offset

    valid_entries = [e for e in toc_entries if e["type"] in ("chapter", "section")]
    skip_pages = set()
    for e in toc_entries:
        if e["type"] == "skip":
            # 标记skip条目对应的页码范围
            skip_pages.add(e["pdf_page"])

    chunks = []
    for i, entry in enumerate(valid_entries):
        start_page = entry["pdf_page"]
        if i + 1 < len(valid_entries):
            end_page = valid_entries[i + 1]["pdf_page"] - 1
        else:
            end_page = max(pages_dict.keys())

        # 检查是否跨进了skip区域
        all_toc = sorted(toc_entries, key=lambda e: e["pdf_page"])
        for te in all_toc:
            if te["type"] == "skip" and start_page < te["pdf_page"] <= end_page:
                end_page = te["pdf_page"] - 1
                break

        text_parts = []
        for pnum in range(start_page, end_page + 1):
            if pnum in pages_dict and pnum not in skip_pages:
                text_parts.append(pages_dict[pnum])

        combined_text = "\n".join(text_parts)
        if len(combined_text.strip()) < 30:
            continue

        chunks.append({
            "chapter": entry["chapter"],
            "section": entry["section"] if entry["type"] == "section" else "",
            "chapter_order": entry.get("order_index", 0),
            "info": f"{entry['chapter']} {entry['section']}".strip(),
            "text": combined_text,
        })

    return chunks


def split_pages_by_regex(pages_dict):
    """后备方案：用正则从正文中检测章节边界"""
    chunks = []
    current_chapter = ""
    current_section = ""
    current_text = ""
    chapter_counter = 0  # 用于生成chapter_order

    skip_keywords = ["跨学科主题学习", "附录", "本书常用地图图例", "活动课", "学史方法", "大事年表"]

    for pnum in sorted(pages_dict.keys()):
        if pnum <= 3:
            continue
        text = pages_dict[pnum]

        should_skip = any(kw in text[:150] for kw in skip_keywords)
        if should_skip:
            if current_text.strip():
                chunks.append({
                    "chapter": current_chapter,
                    "section": current_section,
                    "chapter_order": chapter_counter,
                    "info": f"{current_chapter} {current_section}".strip(),
                    "text": current_text
                })
                current_text = ""
            continue

        # 匹配多种章节格式
        ch = re.search(r"(第[一二三四五六七八九十百]+章|第\s*\d+\s*章)\s+(.+?)(?:\n|$)", text)
        simple_ch = re.search(r"^([一二三四五六七八九十]+)[、\s]+(.{2,30}?)(?:\n|$)", text, re.MULTILINE)
        sec = re.search(r"(第[一二三四五六七八九十百]+节)\s*(.+?)(?:\n|$)", text)
        lesson = re.search(r"(第\s*\d+\s*课)\s*(.+?)(?:\n|$)", text)
        unit = re.search(r"(第[一二三四五六七八九十百]+单元)\s+(.+?)(?:\n|$)", text)
        # 小数点节号（数学/理科教材，如"15.1 分式及其基本性质"），出现在行首
        decimal_sec = re.search(r"^(?:\S+\n)?(\d+\.\d+)\s+([^\n]{2,40})(?:\n|$)", text)

        if sec or lesson or decimal_sec:
            if current_text.strip():
                chunks.append({
                    "chapter": current_chapter,
                    "section": current_section,
                    "chapter_order": chapter_counter,
                    "info": f"{current_chapter} {current_section}".strip(),
                    "text": current_text
                })
            if ch:
                ch_prefix = re.sub(r'\s+', '', ch.group(1))  # 统一去空格：第15章
                current_chapter = f"{ch_prefix} {ch.group(2).strip()}"
            elif unit:
                current_chapter = f"{unit.group(1)} {unit.group(2).strip()}"
            elif simple_ch:
                current_chapter = f"{simple_ch.group(1)} {simple_ch.group(2).strip()}"
            if sec or lesson:
                m = sec or lesson
                current_section = f"{m.group(1).replace(' ','')} {m.group(2).strip()}"
            else:
                # decimal_sec
                current_section = f"{decimal_sec.group(1)} {decimal_sec.group(2).strip()}"
            current_text = text
        elif ch or unit or simple_ch:
            if current_text.strip():
                chunks.append({
                    "chapter": current_chapter,
                    "section": current_section,
                    "chapter_order": chapter_counter,
                    "info": f"{current_chapter} {current_section}".strip(),
                    "text": current_text
                })
            m = ch or unit or simple_ch
            if simple_ch:
                current_chapter = f"{simple_ch.group(1)} {simple_ch.group(2).strip()}"
            elif ch:
                ch_prefix = re.sub(r'\s+', '', ch.group(1))
                current_chapter = f"{ch_prefix} {ch.group(2).strip()}"
            else:
                current_chapter = f"{m.group(1)} {m.group(2).strip()}"
            current_section = ""
            current_text = text
            chapter_counter += 1  # 新章节，计数器+1
        else:
            current_text += "\n" + text

    if current_text.strip():
        chunks.append({
            "chapter": current_chapter,
            "section": current_section,
            "chapter_order": chapter_counter,
            "info": f"{current_chapter} {current_section}".strip(),
            "text": current_text
        })

    return [c for c in chunks if len(c["text"].strip()) > 80]


# ============================================================
# Excel解析
# ============================================================
def _open_xlsx(xlsx_path):
    """打开Excel，返回 (workbook, worksheet, col_map)"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(c).strip() if c else "" for c in row]
        break
    col_map = {h: i for i, h in enumerate(headers)}
    return wb, ws, col_map

def _get_cell(vals, col_map, *names, default=""):
    """按多个候选列名取单元格值，返回第一个非空的"""
    for name in names:
        idx = col_map.get(name)
        if idx is not None and idx < len(vals) and vals[idx] is not None:
            return str(vals[idx]).strip()
    return default


def parse_old_kb(xlsx_path):
    """解析旧库Excel"""
    wb, ws, col_map = _open_xlsx(xlsx_path)

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals or all(v is None for v in vals):
            continue
        g = lambda *names: _get_cell(vals, col_map, *names)

        subject_raw = g("subjectInner", "学科", "subject")
        kid    = g("知识点Kid", "一级知识点id", "Kid")
        kname  = g("知识点名称", "一级知识点")
        sub_title  = g("二级知识点标题")
        sub_detail = g("二级知识点详情")

        if not sub_title and not kname:
            continue

        clean_detail = strip_html(sub_detail)
        records.append({
            "subject": normalize_subject(subject_raw),
            "kid": kid,
            "knowledge_name": kname,
            "sub_title": sub_title,
            "sub_detail": clean_detail,
            "sub_detail_short": clean_detail[:150],
            "chapter": g("篇章名称", "章"),
            "section": g("模块名称", "节"),
            "period": g("学段"),
            "grade": g("年级"),
            "publisher": g("出版社"),
            "volume": g("volume", "册"),
        })

    wb.close()
    return records


def parse_ref_table(xlsx_path):
    """解析参考表格"""
    wb, ws, col_map = _open_xlsx(xlsx_path)

    samples = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(samples) >= 30:
            break
        vals = list(row)
        if not vals or all(v is None for v in vals):
            continue
        g = lambda *names: _get_cell(vals, col_map, *names)
        samples.append({
            "章": g("章", "篇章名称"),
            "节": g("节", "模块名称"),
            "一级知识点": g("一级知识点", "知识点名称"),
            "一级知识点id": g("一级知识点id", "知识点Kid"),
            "二级知识点标题": g("二级知识点标题"),
            "二级知识点详情": strip_html(g("二级知识点详情"))[:200],
            "备注": g("备注"),
        })

    wb.close()
    return samples


def detect_subjects_in_kb(xlsx_path):
    """快速扫描旧库中有哪些学科"""
    wb, ws, col_map = _open_xlsx(xlsx_path)
    subj_idx = col_map.get("subjectInner", col_map.get("学科", col_map.get("subject")))

    subjects = {}
    if subj_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            if subj_idx < len(vals) and vals[subj_idx]:
                norm = normalize_subject(str(vals[subj_idx]).strip())
                subjects[norm] = subjects.get(norm, 0) + 1

    wb.close()
    return sorted(subjects.items(), key=lambda x: -x[1])


def detect_kb_filters(xlsx_path):
    """扫描旧库中所有可筛选维度：学科、学段、年级、出版社、册"""
    wb, ws, col_map = _open_xlsx(xlsx_path)
    dims = {"subjects": {}, "periods": {}, "grades": {}, "publishers": {}, "volumes": {}}
    field_map = {
        "subjects":   col_map.get("subjectInner", col_map.get("学科")),
        "periods":    col_map.get("学段"),
        "grades":     col_map.get("年级"),
        "publishers": col_map.get("出版社"),
        "volumes":    col_map.get("volume", col_map.get("册")),
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        for dim, idx in field_map.items():
            if idx is not None and idx < len(vals) and vals[idx]:
                val = str(vals[idx]).strip()
                if dim == "subjects":
                    val = normalize_subject(val)
                if val:
                    dims[dim][val] = dims[dim].get(val, 0) + 1

    wb.close()
    return {k: sorted(v.items(), key=lambda x: -x[1]) for k, v in dims.items()}


# ============================================================
# 工具函数
# ============================================================
SUBJECT_MAP = {
    "历史": "历史", "中国历史": "历史", "世界历史": "历史",
    "生物": "生物", "生物学": "生物",
    "政治": "政治", "道德与法治": "政治",
    "地理": "地理", "化学": "化学", "物理": "物理", "数学": "数学",
}

def normalize_subject(s):
    if not s:
        return s
    s = s.strip()
    for key, val in SUBJECT_MAP.items():
        if key in s:
            return val
    return s

def strip_html(text):
    if not text:
        return ""
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"</p>", "\n", text, flags=re.I)
    text = re.sub(r"</div>", "\n", text, flags=re.I)
    text = re.sub(r"</li>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = html_mod.unescape(text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()

def text_similarity(a, b):
    if not a or not b:
        return 0.0
    la, lb = len(a), len(b)
    if la == 0 or lb == 0:
        return 0.0
    if max(la, lb) / max(min(la, lb), 1) > 5:
        return 0.1
    return SequenceMatcher(None, a, b).ratio()

def keyword_overlap(a, b):
    if not a or not b:
        return 0.0
    def tokenize(s):
        tokens = re.split(r'[，。、；：\s,.:;!?()（）\[\]【】""''\u00b7\u2014-]+', s)
        return set(t for t in tokens if len(t) >= 2)
    sa, sb = tokenize(a), tokenize(b)
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / max(min(len(sa), len(sb)), 1)

def sanitize_excel_value(val):
    """移除Excel不支持的非法字符（控制字符等）"""
    if not isinstance(val, str):
        return val
    return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]', '', val)

def _make_result(point, note, kb=None, confidence=0):
    """构造标准结果字典，kb为旧库记录（匹配时传入），None表示需新增"""
    return {
        "章": point.get("章", ""),
        "节": point.get("节", ""),
        "chapter_order": point.get("chapter_order", 0),
        "一级知识点": (kb["knowledge_name"] if kb else None) or point.get("一级知识点", ""),
        "一级知识点id": kb["kid"] if kb else "",
        "二级知识点标题": (kb["sub_title"] if kb else None) or point.get("二级知识点标题", ""),
        "二级知识点详情": (kb["sub_detail"] if kb else None) or point.get("二级知识点详情", ""),
        "备注": note,
        "_confidence": confidence,
        "_new_title": point.get("二级知识点标题", ""),
        "_new_detail": point.get("二级知识点详情", ""),
        "_kb_chapter": kb.get("chapter", "") if kb else "",
        "_kb_section": kb.get("section", "") if kb else "",
    }


# ============================================================
# 旧库匹配
# ============================================================
def prefilter_kb_for_chapter(subject_kb, chapter_name, section_name, new_title="", new_k1=""):
    """根据章节名称和知识点标题预筛选相关旧库记录"""
    keywords = set()
    for name in [chapter_name, section_name, new_title, new_k1]:
        tokens = re.split(r'[第一二三四五六七八九十百章节课单元\s——""（）\-的与和及其在从到]+', name)
        keywords.update(t for t in tokens if len(t) >= 2)

    if not keywords and not new_title:
        return subject_kb[:800]

    scored = []
    for r in subject_kb:
        score = 0
        text = f"{r['chapter']} {r['section']} {r['knowledge_name']} {r['sub_title']}"
        for kw in keywords:
            if kw in text:
                score += 1

        # Title containment bonus: if new_title contains KB sub_title or vice versa
        if new_title and r["sub_title"] and len(r["sub_title"]) >= 3:
            if r["sub_title"] in new_title or new_title in r["sub_title"]:
                score += 5  # Strong signal

        scored.append((score, r))

    scored.sort(key=lambda x: -x[0])
    # Take all with score > 0, up to 800
    result = [r for s, r in scored if s > 0][:800]
    # If too few hits, pad with top records
    if len(result) < 200:
        seen = set(id(r) for r in result)
        for _, r in scored:
            if id(r) not in seen:
                result.append(r)
                if len(result) >= 400:
                    break
    return result


def match_with_old_kb(new_points, old_kb_records, api_config=None, log_fn=None, match_config=None):
    if log_fn is None:
        log_fn = lambda msg, t="info": None
    if match_config is None:
        match_config = {}

    threshold = match_config.get("threshold", 70) / 100.0
    w_title = match_config.get("w_title", 50) / 100.0
    w_detail = match_config.get("w_detail", 40) / 100.0
    w_k1 = match_config.get("w_k1", 10) / 100.0
    use_ai = match_config.get("use_ai", True)

    log_fn(f"🔍 代码匹配 ({len(new_points)} 条 vs {len(old_kb_records)} 条旧库)")
    log_fn(f"  参数: 阈值={threshold:.0%} 二级标题={w_title:.0%} 二级详情={w_detail:.0%} 一级知识点={w_k1:.0%}")

    matched_results = []
    unmatched = []

    for point in new_points:
        best_match = None
        best_score = 0

        new_title = point.get("二级知识点标题", "")
        new_detail = point.get("二级知识点详情", "")
        new_k1 = point.get("一级知识点", "")

        candidates = prefilter_kb_for_chapter(
            old_kb_records, point.get("章", ""), point.get("节", ""),
            new_title, new_k1
        )

        for kb in candidates:
            title_sim = text_similarity(new_title, kb["sub_title"])

            # Fast path 1: title nearly identical → high confidence
            if title_sim >= 0.85:
                score = 0.85 + title_sim * 0.15
            # Fast path 2: one title contains the other (e.g. "土地改革" in "土地改革的意义")
            elif (len(new_title) >= 3 and len(kb["sub_title"]) >= 3 and
                  (new_title in kb["sub_title"] or kb["sub_title"] in new_title)):
                score = 0.80 + title_sim * 0.10
            else:
                detail_kw = keyword_overlap(new_detail[:80], kb["sub_detail_short"])
                detail_sim = text_similarity(new_detail[:80], kb["sub_detail_short"])
                detail_score = max(detail_kw, detail_sim)

                k1_sim = text_similarity(new_k1, kb["knowledge_name"])
                score = title_sim * w_title + detail_score * w_detail + k1_sim * w_k1

            if score > best_score:
                best_score = score
                best_match = kb

        if best_score >= threshold and best_match:
            matched_results.append(_make_result(point, "已存在，无需新增", kb=best_match, confidence=round(best_score * 100)))
        else:
            point["_best_score"] = round(best_score, 3)
            unmatched.append(point)

    log_fn(f"✅ 代码匹配完成：匹配 {len(matched_results)} 条，未匹配 {len(unmatched)} 条", "success")

    if unmatched and use_ai and api_config and api_config.get("api_key"):
        log_fn(f"🤖 AI语义比对 {len(unmatched)} 条")
        ai_results = ai_semantic_match(unmatched, old_kb_records, api_config, log_fn)
        matched_results.extend(ai_results)
    else:
        for point in unmatched:
            matched_results.append(_make_result(point, "需新增"))
        if unmatched:
            log_fn(f"⏭️ 跳过AI比对，{len(unmatched)} 条标记为需新增")

    return matched_results


def ai_semantic_match(unmatched_points, old_kb_records, api_config, log_fn):
    results = []
    batch_size = 8

    for i in range(0, len(unmatched_points), batch_size):
        batch = unmatched_points[i:i + batch_size]

        # 为batch预筛选旧库
        relevant_ids = set()
        for pt in batch:
            cands = prefilter_kb_for_chapter(
                old_kb_records, pt.get("章", ""), pt.get("节", ""),
                pt.get("二级知识点标题", ""), pt.get("一级知识点", "")
            )
            for c in cands[:80]:
                relevant_ids.add(id(c))

        kb_list = [r for r in old_kb_records if id(r) in relevant_ids][:150]
        kb_text = "\n".join([
            f"Kid:{r['kid']}|名称:{r['knowledge_name']}|标题:{r['sub_title']}|详情:{r['sub_detail_short']}"
            for r in kb_list
        ])

        batch_json = json.dumps([
            {"index": j, "一级知识点": p.get("一级知识点",""),
             "二级知识点标题": p.get("二级知识点标题",""),
             "二级知识点详情": p.get("二级知识点详情","")[:200]}
            for j, p in enumerate(batch)
        ], ensure_ascii=False)

        prompt = f"""判断以下新教材知识点是否与旧库中已有知识点语义相近。

新教材知识点：
{batch_json}

旧库知识点：
{kb_text}

规则：
1. 比较语义，看考点角度是否一致。允许不同表述但含义相同的匹配
2. 新教材有新考点角度（旧库未涵盖），判为不匹配
3. 匹配到多个只选最贴近的一个

输出JSON数组：
{{"index":0,"matched":true/false,"kid":"匹配的Kid或空","knowledge_name":"","sub_title":"","sub_detail":""}}

只输出JSON数组。"""

        try:
            ai_result = call_llm_api(api_config, prompt)
            parsed = parse_json_response(ai_result)
            parsed_map = {item.get("index", -1): item for item in parsed}

            for j, point in enumerate(batch):
                item = parsed_map.get(j)
                if item and item.get("matched"):
                    ai_kb = {
                        "knowledge_name": item.get("knowledge_name") or point.get("一级知识点", ""),
                        "kid": item.get("kid", ""),
                        "sub_title": item.get("sub_title") or point.get("二级知识点标题", ""),
                        "sub_detail": strip_html(item.get("sub_detail", "")) or point.get("二级知识点详情", ""),
                        "chapter": "", "section": "",
                    }
                    results.append(_make_result(point, "已存在，无需新增", kb=ai_kb, confidence=-1))
                else:
                    results.append(_make_result(point, "需新增"))

            ai_match = sum(1 for it in parsed_map.values() if it.get("matched"))
            log_fn(f"  批次 {i//batch_size+1}: {len(batch)}条 → AI匹配{ai_match}条")

        except Exception as e:
            err_msg = str(e)[:120]
            log_fn(f"  ⚠️ 批次 {i//batch_size+1} 失败: {err_msg}", "error")
            for point in batch:
                results.append(_make_result(point, f"需新增（AI比对失败：{err_msg}）"))

    ai_exist = sum(1 for r in results if "已存在" in r.get("备注", ""))
    log_fn(f"✅ 第二轮完成：AI匹配 {ai_exist} 条，需新增 {len(results)-ai_exist} 条", "success")
    return results


# ============================================================
# LLM API调用（带重试）
# ============================================================
def detect_api_format(api_url):
    """根据API地址自动判断接口格式"""
    url_lower = api_url.lower()
    if "anthropic.com" in url_lower or "/v1/messages" in url_lower:
        return "claude"
    if "generativelanguage.googleapis.com" in url_lower:
        return "gemini"
    # 其余都走OpenAI兼容格式（通义/DeepSeek/智谱/Moonshot/OpenAI等）
    return "openai"


def call_llm_api(api_config, user_prompt, system_prompt=None, max_retries=3):
    url = api_config["api_url"]
    key = api_config["api_key"]
    model = api_config["model"]
    fmt = detect_api_format(url)

    # 根据格式构建请求
    if fmt == "claude":
        body_dict = {
            "model": model,
            "max_tokens": 16000,
            "temperature": 0.1,
            "messages": [{"role": "user", "content": user_prompt}],
        }
        if system_prompt:
            body_dict["system"] = system_prompt
        headers = {
            "Content-Type": "application/json",
            "x-api-key": key,
            "anthropic-version": "2023-06-01",
        }
    elif fmt == "gemini":
        # Gemini REST API: POST .../models/{model}:generateContent?key=KEY
        # 把system和user拼在一起（Gemini的system_instruction格式）
        contents = []
        if system_prompt:
            contents.append({"role": "user", "parts": [{"text": system_prompt + "\n\n" + user_prompt}]})
        else:
            contents.append({"role": "user", "parts": [{"text": user_prompt}]})
        body_dict = {
            "contents": contents,
            "generationConfig": {"temperature": 0.1, "maxOutputTokens": 16000},
        }
        # Gemini的URL格式: .../models/gemini-xxx:generateContent
        if ":generateContent" not in url:
            url = f"{url.rstrip('/')}/models/{model}:generateContent"
        if "?" in url:
            url += f"&key={key}"
        else:
            url += f"?key={key}"
        headers = {"Content-Type": "application/json"}
    else:
        # OpenAI兼容格式
        messages = []
        if system_prompt:
            messages.append({"role": "system", "content": system_prompt})
        messages.append({"role": "user", "content": user_prompt})
        body_dict = {
            "model": model,
            "messages": messages,
            "max_tokens": 16000,
            "temperature": 0.1,
        }
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {key}",
        }

    data = json.dumps(body_dict).encode("utf-8")

    last_error = None
    for attempt in range(max_retries):
        try:
            import urllib.error
            req = urllib.request.Request(url, data=data, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=300) as resp:
                result = json.loads(resp.read().decode("utf-8"))

            # 解析各格式的返回值
            if fmt == "claude":
                # Anthropic: {"content": [{"type":"text","text":"..."}]}
                content = result.get("content", [])
                texts = [c.get("text", "") for c in content if c.get("type") == "text"]
                return "\n".join(texts)
            elif fmt == "gemini":
                # Gemini: {"candidates":[{"content":{"parts":[{"text":"..."}]}}]}
                candidates = result.get("candidates", [])
                if candidates:
                    parts = candidates[0].get("content", {}).get("parts", [])
                    return "\n".join(p.get("text", "") for p in parts)
                return ""
            else:
                # OpenAI兼容
                if "choices" in result and result["choices"]:
                    return result["choices"][0].get("message", {}).get("content", "")
                elif "output" in result:
                    out = result["output"]
                    if isinstance(out, str):
                        return out
                    return (out.get("text", "") or
                            out.get("choices", [{}])[0].get("message", {}).get("content", ""))
                elif "result" in result:
                    return result["result"]
                return ""

        except urllib.error.HTTPError as e:
            # 对不同HTTP错误码提供明确的诊断信息
            error_body = ""
            try:
                error_body = e.read().decode('utf-8', errors='ignore')[:300]
            except:
                pass

            if e.code == 404:
                # 404 不应该重试，直接给出诊断
                if "ark.cn-beijing.volces.com" in url or "ark.cn-shanghai.volces.com" in url or "volces.com" in url:
                    raise ValueError(
                        f"❌ 豆包接入点不存在(HTTP 404)。请检查：\n"
                        f"   1. 接入点ID是否正确（Volcengine控制台 → 方舟 → 推理接入点）\n"
                        f"   2. 接入点所在地域是否与URL地域一致（cn-beijing vs cn-shanghai）\n"
                        f"   3. 该接入点是否被删除或已过期\n"
                        f"   4. API Key是否属于同一账号\n"
                        f"服务器返回: {error_body}"
                    )
                else:
                    raise ValueError(f"API端点不存在(HTTP 404)。请检查API地址和模型名称是否正确。详情: {error_body}")
            elif e.code == 401:
                raise ValueError(f"❌ API Key无效(HTTP 401)。请检查API Key是否正确填写，或是否已过期")
            elif e.code == 403:
                raise ValueError(f"❌ API访问被拒绝(HTTP 403)。请检查API Key的权限或账户是否有可用额度")

            # 其他HTTP错误，等待重试
            last_error = e
            if attempt < max_retries - 1:
                time.sleep((attempt + 1) * 3)

        except Exception as e:
            last_error = e
            if attempt < max_retries - 1:
                time.sleep((attempt + 1) * 3)
    raise last_error


def parse_json_response(text):
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    cleaned = cleaned.strip()

    # 1. Try direct parse
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    # 2. Try extracting complete JSON array
    match = re.search(r"\[[\s\S]*\]", cleaned)
    if match:
        try:
            return json.loads(match.group(0))
        except:
            pass

    # 3. Try extracting single JSON object
    match = re.search(r"\{[\s\S]*\}", cleaned)
    if match:
        try:
            obj = json.loads(match.group(0))
            return [obj] if isinstance(obj, dict) else obj
        except:
            pass

    # 4. Truncated JSON repair: find all complete {...} objects in a broken array
    objects = []
    for m in re.finditer(r'\{[^{}]*\}', cleaned):
        try:
            obj = json.loads(m.group(0))
            if isinstance(obj, dict) and ("二级知识点标题" in obj or "一级知识点" in obj):
                objects.append(obj)
        except:
            pass
    if objects:
        return objects

    raise ValueError(f"无法解析AI返回的JSON: {cleaned[:300]}")


# ============================================================
# 知识点梳理Prompt
# ============================================================
def build_extraction_prompt(ref_samples, chunk):
    ref_text = ""
    if ref_samples:
        ref_text = "\n".join([
            f"  章:{r['章']} | 节:{r['节']} | 一级知识点:{r['一级知识点']} | 二级标题:{r['二级知识点标题']} | 二级详情:{r['二级知识点详情'][:120]}"
            for r in ref_samples[:12]
        ])

    chapter = chunk.get("chapter", "")
    section = chunk.get("section", "")

    system = f"""你是K12教材知识点梳理专家。严格按规则梳理：

## 规则
1. **仅处理正式课文知识性内容**。跳过：活动课、跨学科学习、附录、年表、单元导语、章首引导文字、课后材料、阅读卡片、图片说明、思考题、练习题、"活动"栏目。
2. **章节名称**用我提供的，不要自编。
3. **中等颗粒度**——不要太粗也不要太细。
4. **一级知识点**：归纳分类名称（如"北方地区地形特征""新中国的成立与巩固"）
5. **二级知识点**：一个一级知识点下可以有多个二级知识点。每个二级知识点包含：
   - 标题：用简短、标准化的名词短语（如"开国大典""土地改革""十一届三中全会"）
   - 详情：简洁、规范、教材化纯文本
6. **按主题分类**：相关的知识点归入同一个一级知识点下

{f"## 参考风格{chr(10)}{ref_text}" if ref_text else ""}

## 输出格式
JSON数组，每个元素代表一个一级知识点：
{{"章":"{chapter}","节":"{section}","一级知识点":"...","二级知识点列表":[{{"标题":"...","详情":"..."}},{{"标题":"...","详情":"..."}}]}}

只输出JSON数组。不要markdown代码块。"""

    user = f"""梳理以下教材内容的知识点：

章节：{chapter} {section}

正文：
{chunk['text']}

输出JSON数组。"""

    return system, user


# ============================================================
# 主处理流程
# ============================================================
def process_textbook(task_id, pdf_path, kb_path, ref_path, subject, api_config, log_fn, match_config=None):
    try:
        # Step 1: PDF
        log_fn("📄 第一步：提取PDF文字...", "info")
        pages = extract_pdf_text(pdf_path)
        total_chars = sum(len(t) for t in pages.values())
        log_fn(f"  共 {len(pages)} 页，{total_chars:,} 字符", "success")

        if total_chars < 500:
            log_fn(f"❌ PDF提取文字过少({total_chars}字/共{len(pages)}页)。可能原因：", "error")
            log_fn("   1. 这是扫描版PDF（图片扫描，非电子文字）", "error")
            log_fn("   2. 请确认PDF可以在PDF阅读器中复制文字", "error")
            log_fn("   3. 如需处理扫描版，需先用OCR工具转换", "error")
            return None

        # Step 2: 章节分割
        log_fn("📑 第二步：解析目录、分割章节...", "info")
        toc_entries = parse_toc_entries(pages)

        if toc_entries:
            content_entries = [e for e in toc_entries if e["type"] != "skip"]
            skip_entries = [e for e in toc_entries if e["type"] == "skip"]
            log_fn(f"  目录: {len(content_entries)} 章节, {len(skip_entries)} 跳过项", "info")
            chunks = split_pages_by_toc(pages, toc_entries)
        else:
            log_fn("  未找到目录，用正则分割", "warn")
            chunks = split_pages_by_regex(pages)

        if not chunks:
            log_fn("❌ 未识别到章节内容", "error")
            return None

        for c in chunks:
            log_fn(f"  📌 {c['info'][:50]}  ({len(c['text'])}字)", "info")

        # Step 3: 旧库
        log_fn("🗄️ 第三步：解析旧库表格...", "info")
        all_kb = parse_old_kb(kb_path)
        log_fn(f"  旧库共 {len(all_kb):,} 条", "info")

        # 筛选：学科（必须）
        target_subject = normalize_subject(subject)
        filtered_kb = [r for r in all_kb if r["subject"] == target_subject]
        log_fn(f"  学科「{target_subject}」: {len(filtered_kb):,} 条", "success")

        # 筛选：学段、年级（可选）
        filters = match_config or {}
        for dim, field in [("period","period"),("grade","grade")]:
            fval = filters.get(dim, "")
            if fval and fval != "全部":
                before = len(filtered_kb)
                filtered_kb = [r for r in filtered_kb if r.get(field, "") == fval]
                log_fn(f"  筛选{dim}「{fval}」: {before} → {len(filtered_kb)} 条", "info")

        subject_kb = filtered_kb
        if not subject_kb:
            log_fn(f"  ⚠️ 筛选后无记录，所有知识点将标记为[需新增]", "warn")

        # Step 4: 参考表格
        ref_samples = []
        if ref_path:
            log_fn("📋 解析参考表格...", "info")
            ref_samples = parse_ref_table(ref_path)
            log_fn(f"  {len(ref_samples)} 条样例", "success")

        # Step 5: AI梳理（并行）
        log_fn("🤖 第四步：AI梳理知识点...", "info")
        parallel = match_config.get("parallel", 4) if match_config else 4
        log_fn(f"  共 {len(chunks)} 个章节，{parallel} 路并行处理...", "info")

        from concurrent.futures import ThreadPoolExecutor, as_completed

        def process_one_chunk(idx, chunk):
            system_prompt, user_prompt = build_extraction_prompt(ref_samples, chunk)
            response = call_llm_api(api_config, user_prompt, system_prompt)
            hierarchical_points = parse_json_response(response)

            # 扁平化层级结构
            flat_points = []
            if isinstance(hierarchical_points, list):
                for primary in hierarchical_points:
                    primary_name = primary.get("一级知识点", "")
                    secondary_list = primary.get("二级知识点列表", [])

                    # 如果没有二级知识点列表，尝试兼容旧格式
                    if not secondary_list and primary.get("二级知识点标题"):
                        secondary_list = [{
                            "标题": primary.get("二级知识点标题", ""),
                            "详情": primary.get("二级知识点详情", "")
                        }]

                    for secondary in secondary_list:
                        flat_points.append({
                            "章": chunk["chapter"],  # 强制使用chunk中的章节信息，确保顺序正确
                            "节": chunk["section"],  # 强制使用chunk中的节信息
                            "chapter_order": chunk.get("chapter_order", 0),
                            "一级知识点": primary_name,
                            "二级知识点标题": secondary.get("标题", ""),
                            "二级知识点详情": secondary.get("详情", ""),
                        })
                return idx, flat_points, None
            return idx, [], "格式异常"

        all_new_points = []
        consecutive_fails = 0
        done_count = 0
        failed_early = False
        chunk_results = {}  # idx → points, preserve order

        with ThreadPoolExecutor(max_workers=parallel) as pool:
            futures = {}
            for i, chunk in enumerate(chunks):
                futures[pool.submit(process_one_chunk, i, chunk)] = (i, chunk)

            for future in as_completed(futures):
                i, chunk = futures[future]
                done_count += 1
                try:
                    idx, points, err = future.result()
                    if err:
                        log_fn(f"  [{done_count}/{len(chunks)}] {chunk['info'][:40]} → ⚠️ {err}", "warn")
                    else:
                        chunk_results[idx] = points
                        log_fn(f"  [{done_count}/{len(chunks)}] {chunk['info'][:40]} → ✓ {len(points)} 条", "success")
                        consecutive_fails = 0
                except Exception as e:
                    log_fn(f"  [{done_count}/{len(chunks)}] {chunk['info'][:40]} → ✗ {str(e)[:80]}", "error")
                    consecutive_fails += 1
                    if consecutive_fails >= 3:
                        log_fn("", "error")
                        log_fn("🛑 连续3次失败，自动停止。请检查API配置。", "error")
                        failed_early = True
                        pool.shutdown(wait=False, cancel_futures=True)
                        break

        # Reassemble in original chapter order
        for idx in sorted(chunk_results.keys()):
            all_new_points.extend(chunk_results[idx])

        log_fn(f"  共提取 {len(all_new_points)} 条知识点", "success")

        if not all_new_points:
            log_fn("❌ 未提取到知识点。请检查API配置。", "error")
            return None

        # Step 6: 比对
        log_fn("🔗 第五步：与旧库比对去重...", "info")
        final_results = match_with_old_kb(all_new_points, subject_kb, api_config, log_fn, match_config)

        # Step 7: Excel
        log_fn("📊 第六步：生成Excel...", "info")
        output_name = f"{Path(pdf_path).stem}_知识点梳理_{time.strftime('%m%d_%H%M')}.xlsx"
        output_path = OUTPUT_DIR / output_name
        generate_excel(final_results, str(output_path))

        exist_count = sum(1 for r in final_results if "已存在" in r.get("备注", ""))
        new_count = sum(1 for r in final_results if "需新增" in r.get("备注", ""))
        log_fn(f"🎉 完成！共 {len(final_results)} 条：已存在 {exist_count}，需新增 {new_count}", "success")

        return {
            "filename": output_name,
            "total": len(final_results),
            "exist": exist_count,
            "new_count": new_count,
            "results": final_results,
        }

    except Exception as e:
        log_fn(f"❌ 异常: {traceback.format_exc()}", "error")
        return None


# ============================================================
# Excel生成
# ============================================================
def generate_excel(results, output_path):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # 排序：按章节顺序排列
    def sort_key(r):
        return (
            r.get("chapter_order", 999),
            r.get("章", ""),
            r.get("节", ""),
            r.get("一级知识点", ""),
            r.get("二级知识点标题", "")
        )

    sorted_results = sorted(results, key=sort_key)

    wb = openpyxl.Workbook()

    # ==========================================
    # Sheet 1: 知识点梳理（干净的7列标准格式）
    # ==========================================
    ws = wb.active
    ws.title = "知识点梳理"

    headers = ["章", "节", "一级知识点", "一级知识点id", "二级知识点标题", "二级知识点详情", "备注"]
    hfill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    hfont = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )
    dfont = Font(name="微软雅黑", size=10)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont; c.fill = hfill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    efill = PatternFill(start_color="F0FAF0", end_color="F0FAF0", fill_type="solid")
    nfill = PatternFill(start_color="FFFAF0", end_color="FFFAF0", fill_type="solid")

    for row_idx, r in enumerate(sorted_results, 2):
        is_exist = "已存在" in r.get("备注", "")
        vals = [r.get("章",""), r.get("节",""), r.get("一级知识点",""),
                r.get("一级知识点id",""), r.get("二级知识点标题",""),
                r.get("二级知识点详情",""), r.get("备注","")]
        fill = efill if is_exist else nfill
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=sanitize_excel_value(val))
            c.font = dfont; c.fill = fill; c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    for cl, w in {"A":25,"B":25,"C":20,"D":15,"E":25,"F":50,"G":18}.items():
        ws.column_dimensions[cl].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(sorted_results)+1}"

    # ==========================================
    # Sheet 2: 对比详情（调试和复查用）
    # ==========================================
    ws2 = wb.create_sheet("对比详情")
    dh = ["行号", "备注", "匹配置信度",
          "旧库Kid", "旧库标题", "旧库详情(前80字)", "旧库篇章", "旧库模块",
          "新教材标题", "新教材详情(前80字)"]
    dhfill = PatternFill(start_color="5B4A8A", end_color="5B4A8A", fill_type="solid")
    for col, h in enumerate(dh, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hfont; c.fill = dhfill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, r in enumerate(sorted_results, 2):
        is_exist = "已存在" in r.get("备注", "")
        if not is_exist:
            continue
        conf = r.get("_confidence", 0)
        conf_txt = "AI判定" if conf == -1 else f"{conf}%"
        vals = [
            ri, r.get("备注",""), conf_txt,
            r.get("一级知识点id",""), r.get("二级知识点标题",""),
            (r.get("二级知识点详情","") or "")[:80],
            r.get("_kb_chapter",""), r.get("_kb_section",""),
            r.get("_new_title",""),
            (r.get("_new_detail","") or "")[:80],
        ]
        row_idx = ws2.max_row + 1
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row_idx, column=col, value=sanitize_excel_value(val))
            c.font = dfont; c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    for cl, w in {"A":5,"B":16,"C":10,"D":14,"E":22,"F":35,"G":20,"H":20,"I":22,"J":35}.items():
        ws2.column_dimensions[cl].width = w
    ws2.freeze_panes = "A2"

    # ==========================================
    # Sheet 3: 汇总说明
    # ==========================================
    ws3 = wb.create_sheet("汇总说明")
    ec = sum(1 for r in sorted_results if "已存在" in r.get("备注",""))
    nc = len(sorted_results) - ec
    cm = [r for r in sorted_results if r.get("_confidence",0) > 0]
    am = [r for r in sorted_results if r.get("_confidence",0) == -1]
    hc = sum(1 for r in cm if r["_confidence"] >= 80)
    mc = sum(1 for r in cm if 60 <= r["_confidence"] < 80)
    lc = sum(1 for r in cm if r["_confidence"] < 60)

    sm = [
        ["项目","数量","说明"],
        ["总条数",len(sorted_results),""],
        ["已存在",ec,"匹配到旧库，回填了Kid和内容"],
        ["需新增",nc,"旧库中没有，需要人工补充Kid"],
        ["","",""],
        ["匹配置信度分布","",""],
        ["高置信(>=80%)",hc,"标题高度匹配，可信赖"],
        ["中置信(60-79%)",mc,"有相似性，建议在「对比详情」sheet核实"],
        ["低置信(<60%)",lc,"匹配较勉强，建议核实"],
        ["AI语义匹配",len(am),"代码未匹配，AI判定相似"],
        ["","",""],
        ["说明","",""],
        ["","","「知识点梳理」sheet是干净的7列标准格式，可直接使用"],
        ["","","「对比详情」sheet列出了所有已匹配行的旧库原始信息，用于核实"],
        ["","","如果对匹配结果不满意，可以在网页上调整阈值后重跑"],
    ]

    cst = {}
    for r in sorted_results:
        ch = r.get("章","未知")
        if ch not in cst: cst[ch] = {"t":0,"e":0,"n":0}
        cst[ch]["t"] += 1
        if "已存在" in r.get("备注",""): cst[ch]["e"] += 1
        else: cst[ch]["n"] += 1
    sm.append(["","",""])
    sm.append(["按章统计","总/已存在/需新增",""])
    for ch, st in cst.items():
        sm.append([ch, f"{st['t']}/{st['e']}/{st['n']}", ""])

    for ri, rd in enumerate(sm, 1):
        for col, val in enumerate(rd, 1):
            c = ws3.cell(row=ri, column=col, value=sanitize_excel_value(val))
            if ri == 1 or rd[0] in ("匹配置信度分布","说明","按章统计"):
                c.font = hfont; c.fill = hfill
            else:
                c.font = dfont
            c.border = border
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 55
    wb.save(output_path)


# ============================================================
# 任务管理
# ============================================================
tasks = {}

def run_task(task_id, pdf_path, kb_path, ref_path, subject, api_config, match_config=None):
    task = tasks[task_id]
    task["status"] = "running"
    def log_fn(msg, msg_type="info"):
        task["logs"].append({"time": time.strftime("%H:%M:%S"), "msg": msg, "type": msg_type})
    result = process_textbook(task_id, pdf_path, kb_path, ref_path, subject, api_config, log_fn, match_config)
    task["status"] = "done" if result else "error"
    task["result"] = result


# ============================================================
# HTTP Server
# ============================================================
class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass

    def _json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        path = self.path.split("?")[0]
        if path in ("/", "/index.html"):
            html_file = BASE_DIR / "index.html"
            h = html_file.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(h)))
            self.end_headers()
            self.wfile.write(h)
        elif path.startswith("/api/task/"):
            tid = path.split("/")[-1]
            t = tasks.get(tid)
            if t:
                self._json({"status": t["status"], "logs": t["logs"], "result": t.get("result")})
            else:
                self._json({"error": "任务不存在"}, 404)
        elif path.startswith("/api/download/"):
            fn = urllib.parse.unquote(path[len("/api/download/"):])
            fp = OUTPUT_DIR / fn
            if fp.exists():
                with open(fp, "rb") as f:
                    data = f.read()
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{urllib.parse.quote(fn)}")
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
            else:
                self._json({"error": "文件不存在"}, 404)
        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        path = self.path.split("?")[0]
        if path == "/api/upload":
            self._handle_upload()
        elif path == "/api/start":
            self._handle_start()
        elif path == "/api/detect-subjects":
            self._handle_detect_subjects()
        elif path == "/api/preview-pdf":
            self._handle_preview_pdf()
        else:
            self._json({"error": "未知接口"}, 404)

    def _read_body(self):
        return self.rfile.read(int(self.headers.get("Content-Length", 0)))

    def _handle_upload(self):
        ct = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in ct:
            self._json({"error": "需要multipart/form-data"}, 400)
            return
        content_length = int(self.headers.get("Content-Length", 0))
        max_size = 600 * 1024 * 1024  # 600MB
        if content_length > max_size:
            self._json({"error": f"文件太大（{content_length // 1024 // 1024}MB），最大允许600MB"}, 413)
            return
        from multipart import parse_options_header, MultipartParser
        _, params = parse_options_header(ct)
        boundary = params.get("boundary", "")
        if not boundary:
            self._json({"error": "缺少boundary"}, 400)
            return

        files = {}
        for part in MultipartParser(self.rfile, boundary, content_length):
            name = part.name
            if part.filename:
                safe = f"{uuid.uuid4().hex[:8]}_{part.filename}"
                sp = UPLOAD_DIR / safe
                part.save_as(str(sp))
                files[name] = {"path": str(sp), "name": part.filename, "size": sp.stat().st_size}
            else:
                files[name] = part.value

        self._json({"files": files})

    def _handle_detect_subjects(self):
        body = json.loads(self._read_body())
        kp = body.get("kb_path", "")
        if not kp or not Path(kp).exists():
            self._json({"error": "文件不存在"}, 400)
            return
        try:
            filters = detect_kb_filters(kp)
            # Also return legacy subjects format for compatibility
            self._json({"subjects": filters.get("subjects", []), "filters": filters})
        except Exception as e:
            self._json({"error": str(e)}, 500)

    def _handle_preview_pdf(self):
        body = json.loads(self._read_body())
        pp = body.get("pdf_path", "")
        if not pp or not Path(pp).exists():
            self._json({"error": "文件不存在"}, 400)
            return
        try:
            pages = extract_pdf_text(pp)
            toc = parse_toc_entries(pages)
            chunks = split_pages_by_toc(pages, toc) if toc else split_pages_by_regex(pages)
            self._json({
                "total_pages": len(pages),
                "total_chars": sum(len(t) for t in pages.values()),
                "toc_entries": len(toc),
                "chunks": [{"info": c["info"][:60], "chars": len(c["text"]),
                            "preview": c["text"][:200]} for c in chunks],
            })
        except Exception as e:
            self._json({"error": str(e)}, 500)

    def _handle_start(self):
        body = json.loads(self._read_body())
        pp, kp, rp = body.get("pdf_path"), body.get("kb_path"), body.get("ref_path")
        subj = body.get("subject", "")
        ac = {k: body.get(k, "") for k in ("api_url", "api_key", "model")}
        mc = {
            "threshold": int(body.get("threshold", 70)),
            "w_title": int(body.get("w_title", 50)),
            "w_detail": int(body.get("w_detail", 40)),
            "w_k1": int(body.get("w_k1", 10)),
            "use_ai": body.get("use_ai", True),
            "parallel": int(body.get("parallel", 4)),
            "period": body.get("period", ""),
            "grade": body.get("grade", ""),
        }

        if not pp or not kp:
            self._json({"error": "缺少必要文件"}, 400)
            return
        if not ac["api_key"]:
            self._json({"error": "缺少API Key"}, 400)
            return

        tid = uuid.uuid4().hex[:12]
        tasks[tid] = {"status": "pending", "logs": [], "result": None}
        threading.Thread(target=run_task, args=(tid, pp, kp, rp, subj, ac, mc), daemon=True).start()
        self._json({"task_id": tid})



class ThreadingHTTPServer(ThreadingMixIn, HTTPServer):
    """多线程HTTP服务器，解决浏览器并发请求时卡住的问题"""
    daemon_threads = True

def main():
    server = ThreadingHTTPServer(("0.0.0.0", PORT), Handler)
    print(f"""
╔══════════════════════════════════════════════════╗
║   📚 教材知识点梳理与旧库去重工具 v2.0          ║
╠══════════════════════════════════════════════════╣
║                                                  ║
║   浏览器打开: http://localhost:{PORT}             ║
║                                                  ║
║   工作流:                                        ║
║   1. PDF文字提取 (pdfplumber)                    ║
║   2. 智能目录解析 + 章节分割                     ║
║   3. AI按章节梳理知识点                          ║
║   4. 代码精确匹配 + AI语义去重                   ║
║   5. 输出标准Excel                               ║
║                                                  ║
║   按 Ctrl+C 停止                                 ║
╚══════════════════════════════════════════════════╝
""")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已停止。")
        server.server_close()

if __name__ == "__main__":
    main()
